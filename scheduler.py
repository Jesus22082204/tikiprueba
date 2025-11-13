# scheduler.py - Automatización de recolección y API REST
import schedule
import time
import threading
from datetime import datetime, timedelta
import logging
import json
import os
import sys

# Agregar estos imports al inicio del archivo scheduler.py
from flask import send_file
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Configurar logging
if not os.path.exists('data'):
    os.makedirs('data')

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('data/air_quality.log'),
        logging.StreamHandler()
    ]
)

class DataScheduler:
    def __init__(self, api_key):
        self.collector = AirQualityCollector(api_key)
        self.running = False
    
    def collect_data_job(self):
        """Job que ejecuta la recolección de datos"""
        try:
            logging.info("Iniciando recolección programada de datos")
            successful, failed = self.collector.collect_all_locations()
            logging.info(f"Recolección completada - Exitosos: {successful}, Fallidos: {failed}")
        except Exception as e:
            logging.error(f"Error en recolección programada: {e}")
    
    def start_scheduler(self):
        """Iniciar el scheduler"""
        # Programar recolecciones cada 4 horas
        schedule.every(4).hours.do(self.collect_data_job)
        
        # Recolección diaria a las 6:00 AM
        schedule.every().day.at("06:00").do(self.collect_data_job)
        
        # Recolección diaria a las 18:00 PM
        schedule.every().day.at("18:00").do(self.collect_data_job)
        
        logging.info("Scheduler iniciado - Recolectando cada 4 horas")
        logging.info("Horarios especiales: 06:00 y 18:00")
        
        self.running = True
        
        # Ejecutar primera recolección inmediatamente
        self.collect_data_job()
        
        # Loop principal del scheduler
        while self.running:
            schedule.run_pending()
            time.sleep(60)  # Verificar cada minuto
    
    def stop_scheduler(self):
        """Detener el scheduler"""
        self.running = False
        logging.info("Scheduler detenido")


# API REST usando Flask
from flask import Flask, jsonify, request
from flask_cors import CORS
from database_setup import get_historical_data, get_monthly_statistics
import sqlite3
import calendar

app = Flask(__name__)
CORS(app)  # Permitir requests desde el frontend

@app.route('/api/current/<location_id>')
def get_current_data(location_id):
    """Obtener datos más recientes de una ubicación"""
    try:
        data = get_historical_data(location_id=location_id, limit=1)
        if data:
            return jsonify({'success': True, 'data': data[0]})
        else:
            return jsonify({'success': False, 'error': 'No data found'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/historical/<location_id>')
def get_historical(location_id):
    """Obtener datos históricos de una ubicación"""
    try:
        # Parámetros opcionales
        days = request.args.get('days', 7, type=int)
        limit = request.args.get('limit', 100, type=int)
        
        # Calcular fecha de inicio
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days)
        
        data = get_historical_data(
            location_id=location_id,
            start_date=start_date.isoformat(),
            end_date=end_date.isoformat(),
            limit=limit
        )
        
        return jsonify({'success': True, 'data': data, 'count': len(data)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/monthly-stats/<location_id>/<int:year>/<int:month>')
def get_monthly_stats(location_id, year, month):
    """Obtener estadísticas mensuales para boxplots"""
    try:
        stats = get_monthly_statistics(location_id, year, month)
        if stats:
            return jsonify({'success': True, 'data': stats})
        else:
            return jsonify({'success': False, 'error': 'No data for this month'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/boxplot-data/<location_id>/<int:year>')
def get_boxplot_data(location_id, year):
    """Obtener datos para boxplots de todo el año"""
    try:
        current_month = datetime.now().month
        boxplot_data = []
        
        for month in range(1, current_month):
            stats = get_monthly_statistics(location_id, year, month)
            if stats and stats['count'] >= 10:  # Mínimo 10 datos por mes
                month_name = calendar.month_name[month]
                
                # Calcular percentiles para boxplot (aproximación)
                from database_setup import get_historical_data, get_monthly_statistics, get_connection
                conn = get_connection()
                cursor = conn.cursor()

                ##conn = sqlite3.connect('data/air_quality.db')
                ##cursor = conn.cursor()
                
                # Obtener todos los valores del mes para calcular percentiles
                cursor.execute('''
                SELECT pm2_5, pm10 FROM air_quality_data 
                WHERE location_id = ? 
                AND strftime('%Y', timestamp) = ? 
                AND strftime('%m', timestamp) = ?
                ORDER BY pm2_5
                ''', (location_id, str(year), f"{month:02d}"))
                
                values = cursor.fetchall()
                conn.close()
                
                if values:
                    pm25_values = [v[0] for v in values if v[0] is not None]
                    pm10_values = [v[1] for v in values if v[1] is not None]
                    
                    if pm25_values and pm10_values:
                        pm25_values.sort()
                        pm10_values.sort()
                        
                        n = len(pm25_values)
                        pm25_stats = {
                            'min': round(pm25_values[0], 2),
                            'q1': round(pm25_values[int(n * 0.25)], 2),
                            'median': round(pm25_values[int(n * 0.5)], 2),
                            'q3': round(pm25_values[int(n * 0.75)], 2),
                            'max': round(pm25_values[-1], 2)
                        }
                        
                        n = len(pm10_values)
                        pm10_stats = {
                            'min': round(pm10_values[0], 2),
                            'q1': round(pm10_values[int(n * 0.25)], 2),
                            'median': round(pm10_values[int(n * 0.5)], 2),
                            'q3': round(pm10_values[int(n * 0.75)], 2),
                            'max': round(pm10_values[-1], 2)
                        }
                        
                        boxplot_data.append({
                            'month': month_name,
                            'month_number': month,
                            'pm25': pm25_stats,
                            'pm10': pm10_stats,
                            'data_count': stats['count']
                        })
        
        return jsonify({'success': True, 'data': boxplot_data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/locations')
def get_locations():
    """Obtener todas las ubicaciones disponibles"""
    try:
        from database_setup import get_historical_data, get_monthly_statistics, get_connection
        conn = get_connection()
        cursor = conn.cursor()
        ##conn = sqlite3.connect('data/air_quality.db')
        ##cursor = conn.cursor()
        
        cursor.execute('''
        SELECT DISTINCT location_id, location_name, latitude, longitude,
               COUNT(*) as data_count,
               MAX(timestamp) as last_update
        FROM air_quality_data 
        GROUP BY location_id
        ORDER BY location_name
        ''')
        
        locations = []
        for row in cursor.fetchall():
            locations.append({
                'id': row[0],
                'name': row[1],
                'latitude': row[2],
                'longitude': row[3],
                'data_count': row[4],
                'last_update': row[5]
            })
        
        conn.close()
        return jsonify({'success': True, 'data': locations})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/status')
def get_status():
    """Obtener estado general del sistema"""
    try:
        from database_setup import get_historical_data, get_monthly_statistics, get_connection
        conn = get_connection()
        cursor = conn.cursor()
        
        ##conn = sqlite3.connect('data/air_quality.db')
        ##cursor = conn.cursor()
        
        # Contar total de registros
        cursor.execute('SELECT COUNT(*) FROM air_quality_data')
        total_records = cursor.fetchone()[0]
        
        # Última actualización
        cursor.execute('SELECT MAX(timestamp) FROM air_quality_data')
        last_update = cursor.fetchone()[0]
        
        # Registros por ubicación
        cursor.execute('''
        SELECT location_id, COUNT(*) 
        FROM air_quality_data 
        GROUP BY location_id
        ''')
        location_counts = dict(cursor.fetchall())
        
        conn.close()
        
        return jsonify({
            'success': True,
            'data': {
                'total_records': total_records,
                'last_update': last_update,
                'location_counts': location_counts,
                'database_status': 'active'
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# NUEVO: endpoint para Tendencias y distribución
@app.route('/api/trends/<location_id>')
def get_trends(location_id):
    """Devuelve:
    - pm25_24h: serie de las últimas 24 h (timestamp, valor)
    - pm10_24h: serie de las últimas 24 h (timestamp, valor)
    - aqi_distribution_7d: conteo por categorías 1..5 en los últimos 7 días
    """
    try:
        from database_setup import get_historical_data, get_monthly_statistics, get_connection
        conn = get_connection()
        cursor = conn.cursor()

        ##conn = sqlite3.connect('data/air_quality.db')
        ##cursor = conn.cursor()

        # Serie 24h
        cursor.execute(
            """SELECT timestamp, pm2_5, pm10
                 FROM air_quality_data
                 WHERE location_id = ?
                 AND timestamp >= datetime('now','-1 day')
                 ORDER BY timestamp ASC""",
            (location_id,)
        )
        rows_24h = cursor.fetchall()

        # Distribución AQI 7 días
        cursor.execute(
            """SELECT aqi
                 FROM air_quality_data
                 WHERE location_id = ?
                 AND timestamp >= datetime('now','-7 day')""",
            (location_id,)
        )
        rows_7d = cursor.fetchall()
        conn.close()

        pm25_24h = [{'t': r[0], 'v': round(r[1],2)} for r in rows_24h if r[1] is not None]
        pm10_24h = [{'t': r[0], 'v': round(r[2],2)} for r in rows_24h if r[2] is not None]

        dist = {'1':0, '2':0, '3':0, '4':0, '5':0}
        for (aqi,) in rows_7d:
            if aqi is None: 
                continue
            key = str(int(aqi))
            if key in dist:
                dist[key] += 1

        if len(pm25_24h) < 2 and len(pm10_24h) < 2 and sum(dist.values()) < 1:
            return jsonify({'success': False, 'error': 'not_enough_data'})

        return jsonify({
            'success': True,
            'pm25_24h': pm25_24h,
            'pm10_24h': pm10_24h,
            'aqi_distribution_7d': dist
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    
    # pilas que no se
@app.route('/api/export/<location_id>')
def export_data(location_id):
    """Exportar datos a Excel según el período solicitado"""
    try:
        period = request.args.get('period', '24h')
        
        # Determinar días según el período
        period_days = {
            '24h': 1,
            'month': 30,
            'year': 365
        }
        
        days = period_days.get(period, 1)
        
        # Obtener datos históricos
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days)
        
        data = get_historical_data(
            location_id=location_id,
            start_date=start_date.isoformat(),
            end_date=end_date.isoformat(),
            limit=50000  # Sin límite práctico
        )
        
        if not data:
            return jsonify({'success': False, 'error': 'No hay datos disponibles para exportar'})
        
        # Obtener información de la ubicación
        from database_setup import get_historical_data, get_monthly_statistics, get_connection
        conn = get_connection()
        cursor = conn.cursor()

        ##conn = sqlite3.connect('data/air_quality.db')
        ##cursor = conn.cursor()
        cursor.execute(
            'SELECT location_name, latitude, longitude FROM air_quality_data WHERE location_id = ? LIMIT 1',
            (location_id,)
        )
        location_info = cursor.fetchone()
        conn.close()
        
        location_name = location_info[0] if location_info else location_id
        latitude = location_info[1] if location_info else 'N/A'
        longitude = location_info[2] if location_info else 'N/A'
        
        # Crear DataFrame
        df = pd.DataFrame(data)
        
        # Renombrar columnas para mejor legibilidad
        column_names = {
            'timestamp': 'Fecha y Hora',
            'location_id': 'ID Ubicación',
            'location_name': 'Nombre Ubicación',
            'latitude': 'Latitud',
            'longitude': 'Longitud',
            'aqi': 'Índice de Calidad del Aire (AQI)',
            'pm2_5': 'PM2.5 (µg/m³)',
            'pm10': 'PM10 (µg/m³)',
            'co': 'Monóxido de Carbono - CO (µg/m³)',
            'no': 'Óxido Nítrico - NO (µg/m³)',
            'no2': 'Dióxido de Nitrógeno - NO2 (µg/m³)',
            'o3': 'Ozono - O3 (µg/m³)',
            'so2': 'Dióxido de Azufre - SO2 (µg/m³)',
            'nh3': 'Amoníaco - NH3 (µg/m³)',
            'temperature': 'Temperatura (°C)',
            'humidity': 'Humedad (%)',
            'pressure': 'Presión Atmosférica (hPa)',
            'wind_speed': 'Velocidad del Viento (m/s)'
        }
        
        df = df.rename(columns=column_names)
        
        # Convertir timestamp a formato legible
        if 'Fecha y Hora' in df.columns:
            df['Fecha y Hora'] = pd.to_datetime(df['Fecha y Hora']).dt.strftime('%Y-%m-%d %H:%M:%S')
        
        # Crear archivo Excel en memoria con formato
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Hoja de datos
            df.to_excel(writer, sheet_name='Datos', index=False)
            
            # Obtener el workbook y worksheet
            workbook = writer.book
            worksheet = writer.sheets['Datos']
            
            # Estilos
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Formatear encabezados
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Formatear datos y ajustar ancho de columnas
            for col_idx, column in enumerate(df.columns, 1):
                # Ancho de columna
                max_length = max(
                    df[column].astype(str).apply(len).max(),
                    len(column)
                )
                worksheet.column_dimensions[worksheet.cell(1, col_idx).column_letter].width = min(max_length + 2, 50)
                
                # Bordes para todas las celdas
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row_idx, col_idx)
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')
            
            # Crear hoja de metadatos
            meta_sheet = workbook.create_sheet('Metadatos')
            
            meta_data = [
                ['INFORMACIÓN DEL REPORTE'],
                [''],
                ['Ubicación:', location_name],
                ['ID Ubicación:', location_id],
                ['Latitud:', latitude],
                ['Longitud:', longitude],
                [''],
                ['Período:', period],
                ['Fecha de inicio:', start_date.strftime('%Y-%m-%d %H:%M:%S')],
                ['Fecha de fin:', end_date.strftime('%Y-%m-%d %H:%M:%S')],
                ['Total de registros:', len(df)],
                [''],
                ['Fecha de generación:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                [''],
                ['DESCRIPCIÓN DE CONTAMINANTES'],
                [''],
                ['AQI', 'Índice de Calidad del Aire (1-5): 1=Bueno, 2=Aceptable, 3=Moderado, 4=Deficiente, 5=Muy deficiente'],
                ['PM2.5', 'Material particulado fino (≤2.5 µm). Límite OMS: 15 µg/m³ (24h)'],
                ['PM10', 'Material particulado (≤10 µm). Límite OMS: 45 µg/m³ (24h)'],
                ['CO', 'Monóxido de carbono. Límite OMS: 4 mg/m³ (24h)'],
                ['NO2', 'Dióxido de nitrógeno. Límite OMS: 25 µg/m³ (24h)'],
                ['O3', 'Ozono troposférico. Límite OMS: 100 µg/m³ (8h)'],
                ['SO2', 'Dióxido de azufre. Límite OMS: 40 µg/m³ (24h)'],
                ['NH3', 'Amoníaco'],
            ]
            
            for row in meta_data:
                meta_sheet.append(row)
            
            # Formatear hoja de metadatos
            meta_sheet['A1'].font = Font(bold=True, size=14, color='4472C4')
            meta_sheet['A15'].font = Font(bold=True, size=12, color='4472C4')
            
            # Ajustar ancho de columnas en metadatos
            meta_sheet.column_dimensions['A'].width = 30
            meta_sheet.column_dimensions['B'].width = 80
            
            # Crear hoja de estadísticas
            stats_sheet = workbook.create_sheet('Estadísticas')
            
            # Calcular estadísticas
            numeric_columns = ['PM2.5 (µg/m³)', 'PM10 (µg/m³)', 'Índice de Calidad del Aire (AQI)']
            stats_data = [['Contaminante', 'Promedio', 'Mínimo', 'Máximo', 'Desv. Estándar']]
            
            for col in numeric_columns:
                if col in df.columns:
                    values = pd.to_numeric(df[col], errors='coerce').dropna()
                    if len(values) > 0:
                        stats_data.append([
                            col,
                            round(values.mean(), 2),
                            round(values.min(), 2),
                            round(values.max(), 2),
                            round(values.std(), 2)
                        ])
            
            for row in stats_data:
                stats_sheet.append(row)
            
            # Formatear encabezados de estadísticas
            for cell in stats_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Ajustar ancho de columnas
            for col in range(1, 6):
                stats_sheet.column_dimensions[stats_sheet.cell(1, col).column_letter].width = 25
        
        output.seek(0)
        
        # Nombre del archivo
        period_name = {
            '24h': '24horas',
            'month': 'ultimo_mes',
            'year': 'ultimo_año'
        }.get(period, period)
        
        filename = f"datos_calidad_aire_{location_id}_{period_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logging.error(f"Error al exportar datos: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

    # pilas que no se

##def run_api_server():
    #"""Ejecutar el servidor API"""
    ##app.run(host='127.0.0.1', port=5000, debug=False)

def run_api_server():
    """Ejecutar el servidor API"""
    import os
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port, debug=False)


def main():
    """Función principal para ejecutar scheduler y API"""
    
    # Obtener API key
    api_key = os.getenv('OPENWEATHER_API_KEY')
    if not api_key:
        try:
            with open('config.json', 'r') as f:
                config = json.load(f)
                api_key = config.get('openweather_api_key')
        except FileNotFoundError:
            pass
    
    if not api_key:
        print("❌ API key no configurada. Crea un archivo config.json con tu API key:")
        print('{"openweather_api_key": "tu_api_key_aqui"}')
        return
    
    if len(sys.argv) > 1:
        if sys.argv[1] == 'scheduler':
            # Ejecutar solo el scheduler
            scheduler = DataScheduler(api_key)
            try:
                scheduler.start_scheduler()
            except KeyboardInterrupt:
                scheduler.stop_scheduler()
                
        elif sys.argv[1] == 'api':
            # Ejecutar solo el servidor API
            print("🚀 Iniciando servidor API en http://127.0.0.1:5000")
            run_api_server()
            
        elif sys.argv[1] == 'both':
            # Ejecutar ambos en threads separados
            scheduler = DataScheduler(api_key)
            
            # Thread para el scheduler
            scheduler_thread = threading.Thread(target=scheduler.start_scheduler)
            scheduler_thread.daemon = True
            scheduler_thread.start()
            
            # Thread para el API server
            api_thread = threading.Thread(target=run_api_server)
            api_thread.daemon = True
            api_thread.start()
            
            print("🚀 Sistema completo iniciado:")
            print("   📡 Scheduler: Recolectando datos cada 4 horas")
            print("   🌐 API Server: http://127.0.0.1:5000")
            
            try:
                while True:
                    time.sleep(1)
            except KeyboardInterrupt:
                scheduler.stop_scheduler()
                print("\n✅ Sistema detenido")
    else:
        print("Uso: python scheduler.py [scheduler|api|both]")

if __name__ == "__main__":
    main()
